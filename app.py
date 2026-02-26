import streamlit as st
import pandas as pd
import io
import zipfile
import os
import re # 引入正規表達式來處理複雜的數字字串
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# --- 1. 頁面設定 ---
st.set_page_config(
    page_title="鋼構清單轉換助手",
    page_icon="🏗️",
    layout="wide"
)

# 隱藏預設選單
st.markdown("""
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# --- 側邊欄 ---
with st.sidebar:
    st.header("📖 使用說明")
    st.markdown("""
    1. **拖拉檔案**：將 CSV 拖入右側。
    2. **自動處理**：
       - 自動修正亂碼
       - 自動轉為數值格式
       - 自動排版美化
    3. **下載**：支援單檔或批次下載。
    """)
    st.divider()
    st.caption("Version: 3.0 (Numeric Fix)")

# --- 主畫面 ---
st.title("🏗️ 鋼構構件清單轉換助手")

# --- 核心邏輯 ---

def clean_numeric_data(df):
    """
    強力清洗函數：將所有看起來像數字的文字強制轉為數值類型
    """
    # 1. 先移除欄位名稱的空白
    df.columns = df.columns.str.strip()
    
    # 2. 遍歷所有欄位
    for col in df.columns:
        # 嘗試將該欄位轉換為數值
        # 步驟 A: 先把 NaN 填補為特殊值以免報錯 (可選，這裡先不補)
        
        # 步驟 B: 定義一個轉換函數
        def try_convert(val):
            if pd.isna(val): return val
            if isinstance(val, (int, float)): return val
            
            s = str(val).strip()
            # 移除常見的非數字干擾字元 (千分位逗號, 單位)
            # 例如: "1,234.56" -> "1234.56", "50 Kgs" -> "50"
            s_clean = re.sub(r'[^\d\.\-]', '', s) # 只保留數字、小數點、負號
            
            try:
                # 嘗試轉為浮點數
                num = float(s_clean)
                # 如果是整數 (例如 5.0)，轉為 int 以求美觀
                if num.is_integer():
                    return int(num)
                return num
            except:
                # 轉換失敗 (代表真的是文字，例如 "H200*200")，回傳原始值
                return val

        # 步驟 C: 套用轉換 (Apply)
        # 只有當該欄位大部分看起來都像數字時才轉？不，我們針對每一格嘗試轉換
        # 但為了避免把 "1-2" 這種規格誤轉，我們可以保守一點：
        # 如果該欄位名稱包含 "重"、"量"、"長"、"寬"、"高"、"面積"，則強制嘗試清洗
        keywords = ['數', '量', '重', '長', '寬', '高', '積', '價']
        if any(k in col for k in keywords):
             df[col] = df[col].apply(try_convert)
        else:
             # 對於其他欄位，也嘗試轉，但不強行移除文字 (例如 "123" 轉數字，但 "A123" 不動)
             df[col] = pd.to_numeric(df[col], errors='ignore')
             
    return df

def process_excel_styling(ws, df_len):
    """Excel 美化函數"""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 表頭樣式
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 數據區樣式
    for row in ws.iter_rows(min_row=5, max_row=4 + df_len):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 如果是數字，設定為數值格式 (雖然 Python 轉了，但 Excel 顯示格式也可設定)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.##' # 千分位，最多兩位小數

    # 自動欄寬
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    val_str = str(cell.value)
                    len_val = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    if len_val > max_length: max_length = len_val
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)

def convert_csv_to_excel(file):
    try:
        # --- 讀取與解碼 ---
        content = None
        for enc in ["cp950", "utf-8", "utf-8-sig"]:
            try:
                raw_content = file.getvalue().decode(enc)
                content = raw_content.splitlines()
                break
            except UnicodeDecodeError: continue
        
        if content is None: return None, None, "❌ 無法識別檔案編碼"

        # --- 表頭偵測 ---
        header_idx = -1
        exclude_keywords = ["案號", "案名", "頁碼", "日期", "統計表", "清單", "次加總", "Total"]
        required_keywords = ["編號", "規格", "材質", "長度", "重量", "單重", "單量", "數量"]
        
        for i, line in enumerate(content[:20]):
            if "," not in line: continue
            if any(k in line for k in exclude_keywords): continue
            if any(k in line for k in required_keywords):
                header_idx = i; break
        
        if header_idx == -1:
            for i, line in enumerate(content[:20]):
                if "," in line and any(k in line for k in required_keywords):
                    header_idx = i; break
        
        if header_idx == -1: return None, None, "❌ 找不到有效表頭"

        # --- 提取 Metadata ---
        meta = {}
        for line in content[:header_idx]:
            clean_line = line.replace(",", " ").replace(":", " ")
            parts = clean_line.split()
            if "案號" in line:
                for p in parts:
                    if p.isdigit() and len(p) > 3: meta["案號"] = p
            if "頁碼" in line:
                try: meta["頁碼"] = line.split("頁碼")[1].strip(":,").strip()
                except: pass
            if "案名" in line:
                try:
                    start = line.find("案名") + 2
                    end = line.find("日期") if "日期" in line else len(line)
                    meta["案名"] = line[start:end].strip(":, ").strip()
                except: pass
            if "日期" in line:
                try: meta["日期"] = line.split("日期")[1].strip(":, ").strip()
                except: pass

        # --- 數據標準化 ---
        data_lines = content[header_idx:]
        max_cols = 0
        for line in data_lines: max_cols = max(max_cols, line.count(",") + 1)
        
        header_line = data_lines[0]
        current_cols = header_line.count(",") + 1
        if current_cols < max_cols:
            header_line += "," * (max_cols - current_cols)
            data_lines[0] = header_line
            
        final_csv_str = "\n".join(data_lines)
        
        # --- Pandas 處理 ---
        df = pd.read_csv(io.StringIO(final_csv_str))
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        df = df.dropna(how='all')

        # [NEW] 強制數值轉型
        df = clean_numeric_data(df)

        # --- 寫入 Excel ---
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=3)
            ws = writer.sheets['Sheet1']
            
            font_meta = Font(bold=True)
            ws['A1'] = f"案號: {meta.get('案號', '')}"; ws['A1'].font = font_meta
            ws['D1'] = f"頁碼: {meta.get('頁碼', '')}"
            ws['A2'] = f"案名: {meta.get('案名', '')}"; ws['A2'].font = font_meta
            ws['D2'] = f"日期: {meta.get('日期', '')}"
            
            process_excel_styling(ws, len(df))

        output.seek(0)
        return output, df, None

    except Exception as e:
        return None, None, f"解析錯誤: {str(e)}"

# --- 介面互動區 ---
uploaded_files = st.file_uploader("📥 請上傳檔案", type="csv", accept_multiple_files=True)

if uploaded_files:
    st.divider()
    if len(uploaded_files) == 1:
        file = uploaded_files[0]
        excel_data, df, error = convert_csv_to_excel(file)
        if error: st.error(error)
        else:
            col1, col2 = st.columns([1, 2])
            with col1:
                st.success("✅ 轉換成功！")
                st.metric(label="資料筆數", value=len(df))
                new_filename = os.path.splitext(file.name)[0] + ".xlsx"
                st.download_button("📥 下載 Excel", excel_data, new_filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
            with col2:
                st.write("📊 **數據預覽 (已轉為數值)**")
                st.dataframe(df.head(8), use_container_width=True)
    else:
        st.info(f"⚡ 正在批次處理 {len(uploaded_files)} 個檔案...")
        zip_buffer = io.BytesIO()
        success_count = 0
        failed_log = []
        progress_bar = st.progress(0)
        
        with zipfile.ZipFile(zip_buffer, "w") as zf:
            for i, file in enumerate(uploaded_files):
                excel_data, _, error = convert_csv_to_excel(file)
                if excel_data:
                    fname = os.path.splitext(file.name)[0] + ".xlsx"
                    zf.writestr(fname, excel_data.getvalue())
                    success_count += 1
                else: failed_log.append(f"{file.name}: {error}")
                progress_bar.progress((i + 1) / len(uploaded_files))
        
        zip_buffer.seek(0)
        if success_count > 0:
            st.success(f"🎉 成功轉換 {success_count} 個檔案")
            st.download_button(f"📦 下載壓縮檔", zip_buffer, "converted_steel_lists.zip", "application/zip", type="primary", use_container_width=True)
        if failed_log:
            with st.expander("查看失敗原因"):
                for log in failed_log: st.write(log)
else:
    st.info("👆 請從上方上傳 CSV 檔案以開始使用")